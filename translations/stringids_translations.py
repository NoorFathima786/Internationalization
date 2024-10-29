import os
import openpyxl
from xml.etree import ElementTree as ET

def update_ts_files(excel_file, ts_files_directory):
    # Load Excel workbook
    wb = openpyxl.load_workbook(excel_file)
    MessageStrings = wb["MessageStrings"]

    # Get list of languages (column headers except for the first one)
    languages = [cell.value.strip() for cell in MessageStrings[1][1:] if cell.value]

    # Get list of language tags (second row)
    language_tags = [cell.value.strip() for cell in MessageStrings[2][1:] if cell.value]

    # Extract all source strings from the Excel file
    all_sources = set()
    for row in MessageStrings.iter_rows(min_row=3, values_only=True):
        if row[0] is not None and row[0].startswith("string_id_") and ' ' not in row[0]:
            all_sources.add(row[0].lower())

    # Iterate over each language column
    for language_index, (language, lang_tag) in enumerate(zip(languages, language_tags), start=1):
        print(f"Processing language: '{language}' with tag '{lang_tag}'")  # Debug output

        # Extract translations for the current language
        translations = {}
        for row in MessageStrings.iter_rows(min_row=3, values_only=True):
            source = row[0]
            if len(row) > language_index:
                translation = row[language_index]
            else:
                translation = None

            if source and translation:
                translations[source] = translation

        # Update .ts file for the current language
        ts_file = os.path.join(ts_files_directory, f"{language}.ts")
        if os.path.exists(ts_file):
            # Parse the .ts file while preserving the DOCTYPE and <name> tags
            with open(ts_file, 'r', encoding='utf-8') as file:
                ts_content = file.read()
            tree = ET.ElementTree(ET.fromstring(ts_content))
            root = tree.getroot()

            # Update the language attribute in the <TS> tag
            root.set('language', lang_tag)

            # Collect source strings from the .ts file
            ts_sources = set()
            for context in root.findall('context'):
                for message in context.findall('message'):
                    source = message.find('source').text
                    ts_sources.add(source)
                    if source in translations:
                        translation_element = message.find('translation')
                        if translation_element is not None:
                            translation_element.text = translations[source]

            # Identify and print missing source strings
            missing_sources = ts_sources - all_sources
            if missing_sources:
                print(f"Missing sources in Excel for language '{language}':")
                for missing_source in missing_sources:
                    print(missing_source)

            # Write the updated tree back to the file, preserving the DOCTYPE and <name> tags
            xml_str = ET.tostring(root, encoding='utf-8', method='xml').decode('utf-8')
            # Fix for preserving the name tag formatting
            xml_str = xml_str.replace('<name />', '<name></name>')
            xml_str = f'<?xml version="1.0" encoding="utf-8"?>\n<!DOCTYPE TS>\n{xml_str}'
            with open(ts_file, 'w', encoding='utf-8') as file:
                file.write(xml_str)
        else:
            print(f"Warning: .ts file for language '{language}' not found at '{ts_file}'.")

def main():
    # Path to Excel file containing translations
    excel_file = "translations.xlsx"

    # Directory containing .ts files
    ts_files_directory = "."

    update_ts_files(excel_file, ts_files_directory)

if __name__ == "__main__":
    main()
